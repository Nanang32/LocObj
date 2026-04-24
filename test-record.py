import cv2
import numpy as np
import time
import urllib.request
import os
from collections import OrderedDict
from ultralytics import RTDETR
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
#  KONFIGURASI
# ─────────────────────────────────────────────
PERSON_CLASS_ID  = 0
CONF_THRESHOLD   = 0.45
FACE_RATIO_TOP   = 0.28
FACE_CONFIRM_MIN = 0.0
W_IOU = 0.50; W_DIST = 0.30; W_APP = 0.20
MIN_MATCH = 0.30; GRACE_SEC = 9.9; MAX_DIST = 300

PROTO_PATH = "deploy.prototxt"
MODEL_PATH = "face_detector.caffemodel"
PROTO_URL  = "https://raw.githubusercontent.com/opencv/opencv/master/samples/dnn/face_detector/deploy.prototxt"
MODEL_URL  = "https://github.com/opencv/opencv_3rdparty/raw/dnn_samples_face_detector_20170830/res10_300x300_ssd_iter_140000.caffemodel"

# Threshold anomali
ANOMALY_SPEED_PX_PER_SEC = 300
ANOMALY_CONF_DROP        = 0.25
ANOMALY_SIZE_RATIO       = 2.5

# ─────────────────────────────────────────────
#  ANSI TERMINAL
# ─────────────────────────────────────────────
AN = {'rst':'\033[0m','b':'\033[1m','cy':'\033[96m','gr':'\033[92m',
      'yw':'\033[93m','rd':'\033[91m','gy':'\033[90m','wh':'\033[97m'}
def cp(t,*s): print(''.join(AN.get(x,'') for x in s)+t+AN['rst'])
def sep(): cp("─"*64,'gy')
def ts(): return time.strftime('%H:%M:%S')

def fmt(sec):
    if sec is None: return "-"
    if sec < 60: return f"{sec:.3f}s"
    m = int(sec // 60); return f"{m}m {sec - m*60:.1f}s"

def format_dur(sec):
    if sec is None: return "-"
    if sec < 60: return f"{sec:06.3f}s"
    m = int(sec // 60); return f"{m}:{sec - m*60:06.3f}s"

# ─────────────────────────────────────────────
#  EXCEL RECORDER
# ─────────────────────────────────────────────
class ExcelRecorder:
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ACT_FILLS = {
        'MASUK'    : PatternFill("solid", fgColor="E2EFDA"),
        'HILANG'   : PatternFill("solid", fgColor="FFF2CC"),
        'KEMBALI'  : PatternFill("solid", fgColor="DDEBF7"),
        'EXPIRE'   : PatternFill("solid", fgColor="FCE4D6"),
        'PROMOSI'  : PatternFill("solid", fgColor="EAD1DC"),
        'ANOMALI'  : PatternFill("solid", fgColor="FF9999"),
        'TERMINATE': PatternFill("solid", fgColor="D9D9D9"),
    }
    THIN = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def __init__(self):
        self.filename = (f"face_tracker_"
                         f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        self.wb = openpyxl.Workbook()

        self.ws_act = self.wb.active
        self.ws_act.title = "Aktivitas ID"
        self._init_activity_sheet()

        self.ws_sum = self.wb.create_sheet("Ringkasan ID")
        self._init_summary_sheet()

        self.ws_ano = self.wb.create_sheet("Anomali")
        self._init_anomaly_sheet()

        self.act_row = 2
        self.ano_row = 2
        self.sum_row = 2
        self._summary = {}
        self._save()
        cp(f"  Excel: {self.filename}", 'cy', 'b')

    def _write_header(self, ws, cols):
        for ci, (title, width) in enumerate(cols, start=1):
            c = ws.cell(row=1, column=ci, value=title)
            c.fill = self.HDR_FILL
            c.font = self.HDR_FONT
            c.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
            c.border = self.THIN
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def _init_activity_sheet(self):
        self._write_header(self.ws_act, [
            ("No", 6), ("Timestamp", 22), ("Waktu", 10), ("Event", 12),
            ("Display ID", 10), ("Internal ID", 12), ("Conf", 8),
            ("In-Frame Sejak", 18), ("Durasi ON", 14), ("Durasi OFF", 14),
            ("Kembali?", 10), ("ID 1 Sekarang", 14), ("Catatan", 45),
        ])

    def _init_summary_sheet(self):
        self._write_header(self.ws_sum, [
            ("Display ID Awal", 14), ("Internal ID", 12),
            ("Conf Rata²", 10), ("Pertama Masuk", 20),
            ("Terakhir Terlihat", 20), ("Total ON", 14),
            ("Total OFF", 14), ("Jml Kembali", 12),
            ("Jml Anomali", 12), ("Status Akhir", 16), ("Catatan", 30),
        ])

    def _init_anomaly_sheet(self):
        self._write_header(self.ws_ano, [
            ("No", 6), ("Timestamp", 22), ("Waktu", 10),
            ("Display ID", 10), ("Internal ID", 12),
            ("Tipe Anomali", 22), ("Nilai", 14),
            ("Threshold", 14), ("Deskripsi", 55),
        ])

    # ── Tulis event aktivitas ─────────────────
    def record_event(self, event_type, display_id, internal_id, conf,
                     enter_time, on_dur, off_dur, did_kembali,
                     id1_now, catatan=""):
        now_dt = datetime.now()
        fill   = self.ACT_FILLS.get(event_type, PatternFill())
        vals = [
            self.act_row - 1,
            now_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3],
            now_dt.strftime('%H:%M:%S'),
            event_type,
            f"ID {display_id}",
            internal_id,
            round(conf, 3) if isinstance(conf, float) else "-",
            (datetime.fromtimestamp(enter_time).strftime('%H:%M:%S')
             if enter_time else "-"),
            fmt(on_dur),
            fmt(off_dur),
            "Ya" if did_kembali else "Tidak",
            f"ID {id1_now}" if id1_now else "-",
            catatan,
        ]
        for ci, v in enumerate(vals, start=1):
            c = self.ws_act.cell(row=self.act_row, column=ci, value=v)
            c.fill   = fill
            c.border = self.THIN
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if display_id == 1 and ci == 5:
                c.font = Font(bold=True, color="1F3864")
        self.ws_act.row_dimensions[self.act_row].height = 18
        self.act_row += 1
        self._save()

    # ── Tulis anomali ─────────────────────────
    def record_anomaly(self, display_id, internal_id, atype,
                       value, threshold, desc):
        now_dt = datetime.now()
        vals = [
            self.ano_row - 1,
            now_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3],
            now_dt.strftime('%H:%M:%S'),
            f"ID {display_id}",
            internal_id,
            atype,
            round(value, 3),
            threshold,
            desc,
        ]
        for ci, v in enumerate(vals, start=1):
            c = self.ws_ano.cell(row=self.ano_row, column=ci, value=v)
            c.fill   = self.ACT_FILLS['ANOMALI']
            c.font   = Font(bold=(ci in (4, 6)), color="7B0000")
            c.border = self.THIN
            c.alignment = Alignment(vertical='center', wrap_text=True)
        self.ws_ano.row_dimensions[self.ano_row].height = 18
        self.ano_row += 1
        self._save()

    # ── Update buffer ringkasan ───────────────
    def update_summary_buffer(self, iid, display_id, conf,
                               enter_time, on_dur, total_off,
                               n_kembali, n_anomali):
        prev = self._summary.get(iid, {})
        self._summary[iid] = {
            'display_id' : display_id,
            'conf_list'  : prev.get('conf_list', []) + [conf],
            'enter_time' : enter_time,
            'on_dur'     : on_dur,
            'total_off'  : total_off,
            'n_kembali'  : n_kembali,
            'n_anomali'  : n_anomali,
            'last_seen'  : time.time(),
            'status'     : prev.get('status', 'aktif'),
        }

    def mark_expired(self, iid):
        if iid in self._summary:
            self._summary[iid]['status']    = 'expire'
            self._summary[iid]['last_seen'] = time.time()

    # ── Tulis ringkasan saat terminate ───────
    def write_summary(self, tracks):
        now = time.time()
        for iid, t in tracks.items():
            if iid in self._summary:
                self._summary[iid]['status']    = 'aktif_saat_berhenti'
                self._summary[iid]['last_seen'] = now

        for iid, s in sorted(self._summary.items(),
                              key=lambda x: x[1].get('enter_time', 0)):
            conf_avg = (sum(s['conf_list']) / len(s['conf_list'])
                        if s['conf_list'] else 0)
            status   = s.get('status', '-')
            fill     = (self.ACT_FILLS['EXPIRE']
                        if status == 'expire'
                        else self.ACT_FILLS['MASUK']
                        if status == 'aktif_saat_berhenti'
                        else PatternFill())
            vals = [
                f"ID {s['display_id']}",
                iid,
                round(conf_avg, 3),
                (datetime.fromtimestamp(s['enter_time'])
                 .strftime('%Y-%m-%d %H:%M:%S')
                 if s.get('enter_time') else "-"),
                (datetime.fromtimestamp(s['last_seen'])
                 .strftime('%Y-%m-%d %H:%M:%S')
                 if s.get('last_seen') else "-"),
                fmt(s.get('on_dur', 0)),
                fmt(s.get('total_off', 0)),
                s.get('n_kembali', 0),
                s.get('n_anomali', 0),
                status,
                "",
            ]
            for ci, v in enumerate(vals, start=1):
                c = self.ws_sum.cell(row=self.sum_row, column=ci, value=v)
                c.fill   = fill
                c.border = self.THIN
                c.alignment = Alignment(vertical='center')
            self.ws_sum.row_dimensions[self.sum_row].height = 18
            self.sum_row += 1
        self._save()

    # ── Baris TERMINATE ───────────────────────
    def record_terminate(self, total_ids, duration_sec):
        now_dt = datetime.now()
        vals = [
            self.act_row - 1,
            now_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3],
            now_dt.strftime('%H:%M:%S'),
            "TERMINATE", "-", "-", "-", "-", "-", "-", "-", "-",
            (f"Sistem berhenti. Total ID unik: {total_ids}. "
             f"Durasi sesi: {fmt(duration_sec)}"),
        ]
        for ci, v in enumerate(vals, start=1):
            c = self.ws_act.cell(row=self.act_row, column=ci, value=v)
            c.fill   = self.ACT_FILLS['TERMINATE']
            c.font   = Font(bold=True, italic=True)
            c.border = self.THIN
            c.alignment = Alignment(vertical='center', wrap_text=True)
        self.act_row += 1
        self._save()
        cp(f"  Excel tersimpan: {self.filename}", 'gr', 'b')

    def _save(self):
        try:
            self.wb.save(self.filename)
        except Exception as e:
            cp(f"  [WARN] Gagal simpan Excel: {e}", 'yw')


# ─────────────────────────────────────────────
#  ANOMALY DETECTOR
# ─────────────────────────────────────────────
class AnomalyDetector:
    def __init__(self):
        self._prev = {}   # iid -> {cx, cy, area, conf, time}

    def check(self, iid, display_id, bbox, conf, recorder):
        """
        bbox = t['bbox']  ← key yang benar di dalam tracks dict
        """
        now  = time.time()
        # ── Pastikan bbox valid sebelum diproses
        if bbox is None or len(bbox) < 4:
            return []

        x1, y1, x2, y2 = [int(v) for v in bbox]
        cx   = (x1 + x2) // 2
        cy   = (y1 + y2) // 2
        area = max(1, (x2 - x1) * (y2 - y1))
        anomalies = []

        if iid in self._prev:
            prev = self._prev[iid]
            dt   = now - prev['time']
            if dt > 0:
                # 1. Anomali kecepatan
                dist  = np.hypot(cx - prev['cx'], cy - prev['cy'])
                speed = dist / dt
                if speed > ANOMALY_SPEED_PX_PER_SEC:
                    desc = (f"Centroid {dist:.1f}px dalam {dt:.3f}s "
                            f"= {speed:.1f}px/s (threshold "
                            f"{ANOMALY_SPEED_PX_PER_SEC}px/s)")
                    recorder.record_anomaly(
                        display_id, iid, "KECEPATAN_TINGGI",
                        speed, ANOMALY_SPEED_PX_PER_SEC, desc)
                    anomalies.append(("KECEPATAN", speed))
                    cp(f"  [ANOMALI] ID {display_id}  "
                       f"kecepatan {speed:.0f}px/s", 'rd', 'b')

                # 2. Anomali confidence drop
                conf_drop = prev['conf'] - conf
                if conf_drop > ANOMALY_CONF_DROP:
                    desc = (f"Conf turun {prev['conf']:.2f}→{conf:.2f} "
                            f"(drop {conf_drop:.2f} > {ANOMALY_CONF_DROP})")
                    recorder.record_anomaly(
                        display_id, iid, "CONFIDENCE_DROP",
                        conf_drop, ANOMALY_CONF_DROP, desc)
                    anomalies.append(("CONF_DROP", conf_drop))

                # 3. Anomali perubahan ukuran bbox
                size_ratio = max(area / prev['area'],
                                 prev['area'] / area)
                if size_ratio > ANOMALY_SIZE_RATIO:
                    desc = (f"Ukuran {size_ratio:.1f}x berubah "
                            f"({prev['area']}→{area}px², "
                            f"threshold {ANOMALY_SIZE_RATIO}x)")
                    recorder.record_anomaly(
                        display_id, iid, "PERUBAHAN_UKURAN",
                        size_ratio, ANOMALY_SIZE_RATIO, desc)
                    anomalies.append(("SIZE", size_ratio))

        self._prev[iid] = {
            'cx': cx, 'cy': cy, 'area': area,
            'conf': conf, 'time': now,
        }
        return anomalies

    def remove(self, iid):
        self._prev.pop(iid, None)


# ─────────────────────────────────────────────
#  DOWNLOAD + FACE VALIDATOR
# ─────────────────────────────────────────────
def download_face_model():
    if not os.path.exists(PROTO_PATH):
        print("Mengunduh prototxt...")
        urllib.request.urlretrieve(PROTO_URL, PROTO_PATH)
    if not os.path.exists(MODEL_PATH):
        print("Mengunduh caffemodel...")
        urllib.request.urlretrieve(MODEL_URL, MODEL_PATH)


class FaceValidator:
    def __init__(self):
        download_face_model()
        self.net = cv2.dnn.readNetFromCaffe(PROTO_PATH, MODEL_PATH)
        cp("Face DNN validator siap.", 'gr')

    def has_face(self, frame, bbox, min_confidence=0.5):
        x1, y1, x2, y2 = bbox
        face_y2 = min(y1 + int((y2 - y1) * FACE_RATIO_TOP), frame.shape[0])
        fx1, fy1 = max(x1, 0), max(y1, 0)
        fx2 = min(x2, frame.shape[1])
        roi = frame[fy1:face_y2, fx1:fx2]
        if roi.shape[0] < 20 or roi.shape[1] < 20:
            return True, [x1, y1, x2, face_y2]
        blob = cv2.dnn.blobFromImage(
            cv2.resize(roi, (300, 300)), 1.0,
            (300, 300), (104.0, 177.0, 123.0))
        self.net.setInput(blob)
        dets = self.net.forward()
        bc, bb = 0.0, None
        rh, rw = roi.shape[:2]
        for i in range(dets.shape[2]):
            c = float(dets[0, 0, i, 2])
            if c > bc:
                bc = c
                bb = [int(dets[0,0,i,3]*rw)+fx1,
                      int(dets[0,0,i,4]*rh)+fy1,
                      int(dets[0,0,i,5]*rw)+fx1,
                      int(dets[0,0,i,6]*rh)+fy1]
        if bc >= min_confidence and bb:
            return True, bb
        if FACE_CONFIRM_MIN == 0.0:
            return True, [x1, y1, x2, face_y2]
        return False, None


# ─────────────────────────────────────────────
#  UTILITAS TRACKING
# ─────────────────────────────────────────────
def compute_iou(b1, b2):
    ix1=max(b1[0],b2[0]); iy1=max(b1[1],b2[1])
    ix2=min(b1[2],b2[2]); iy2=min(b1[3],b2[3])
    inter=max(0,ix2-ix1)*max(0,iy2-iy1)
    u=(b1[2]-b1[0])*(b1[3]-b1[1])+(b2[2]-b2[0])*(b2[3]-b2[1])-inter
    return inter/u if u>0 else 0.0

def centroid(b):
    return ((b[0]+b[2])//2, (b[1]+b[3])//2)

def ndist(b1, b2):
    c1=centroid(b1); c2=centroid(b2)
    return min(np.hypot(c1[0]-c2[0], c1[1]-c2[1])/MAX_DIST, 1.0)

def get_app(frame, bbox):
    x1,y1,x2,y2=[max(0,int(v)) for v in bbox]
    roi=frame[y1:y2,x1:x2]
    if roi.size==0: return np.zeros(48)
    hsv=cv2.cvtColor(roi,cv2.COLOR_BGR2HSV)
    h=cv2.calcHist([hsv],[0],None,[16],[0,180]).flatten()
    s=cv2.calcHist([hsv],[1],None,[16],[0,256]).flatten()
    v=cv2.calcHist([hsv],[2],None,[16],[0,256]).flatten()
    hist=np.concatenate([h,s,v]).astype(np.float32)
    n=hist.sum(); return hist/n if n>0 else hist

def app_sim(h1, h2):
    d=np.dot(h1,h2); n=np.linalg.norm(h1)*np.linalg.norm(h2)
    return float(d/n) if n>0 else 0.0

def mscore(track, bbox, app):
    return (W_IOU * compute_iou(track['bbox'], bbox)
           +W_DIST * (1 - ndist(track['bbox'], bbox))
           +W_APP  * app_sim(track['appearance'], app))


# ─────────────────────────────────────────────
#  ID MANAGER
#  Semua track menyimpan bbox wajah sebagai 'bbox'
#  (bukan 'face_bbox') agar konsisten di seluruh kode
# ─────────────────────────────────────────────
class IDManager:
    def __init__(self):
        self.tracks      = OrderedDict()
        self._id_counter = 0

    def _rank(self):
        for rank, (_, t) in enumerate(self.tracks.items(), start=1):
            t['display_id'] = rank

    def update(self, frame, detections):
        """
        detections: list of dict dengan key:
          'face_bbox', 'person_bbox', 'conf'
        Track disimpan dengan key 'bbox' (= face_bbox) dan 'person_bbox'.
        """
        now = time.time()
        for d in detections:
            # Appearance dihitung dari face_bbox
            d['appearance'] = get_app(frame, d['face_bbox'])

        # Greedy matching
        used, matched = {}, {}
        for tid, track in self.tracks.items():
            bs, bi = -1, -1
            for i, d in enumerate(detections):
                if i in used: continue
                s = mscore(track, d['face_bbox'], d['appearance'])
                if s > bs: bs, bi = s, i
            if bs >= MIN_MATCH:
                matched[tid] = bi; used[bi] = True

        # Update track yang cocok
        for tid, di in matched.items():
            d = detections[di]; t = self.tracks[tid]
            # Akumulasi off-time jika baru kembali
            if t['missing'] and t.get('off_start'):
                t['total_off'] = t.get('total_off', 0.0) + (now - t['off_start'])
                t['off_start'] = None
                t['n_kembali'] = t.get('n_kembali', 0) + 1
            # ── Simpan sebagai 'bbox' bukan 'face_bbox' ──
            t['bbox']        = d['face_bbox']   # ← KEY KONSISTEN
            t['person_bbox'] = d['person_bbox']
            t['conf']        = d['conf']
            t['appearance']  = d['appearance']
            t['last_seen']   = now
            t['missing']     = False

        # Tandai hilang
        for tid in self.tracks:
            if tid not in matched:
                t = self.tracks[tid]
                if not t['missing']:
                    t['off_start'] = now
                t['missing'] = True

        # Track baru — append ke akhir OrderedDict
        for i, d in enumerate(detections):
            if i not in used:
                self.tracks[self._id_counter] = {
                    'bbox':        d['face_bbox'],   # ← KEY KONSISTEN
                    'person_bbox': d['person_bbox'],
                    'conf':        d['conf'],
                    'appearance':  d['appearance'],
                    'enter_time':  now,
                    'last_seen':   now,
                    'missing':     False,
                    'display_id':  None,
                    'off_start':   None,
                    'total_off':   0.0,
                    'n_kembali':   0,
                    'n_anomali':   0,
                }
                self._id_counter += 1

        self._rank()

        # Kumpulkan expired (sudah ada display_id) sebelum dihapus
        expired = []
        for k, t in self.tracks.items():
            if t['missing'] and (now - t['last_seen']) > GRACE_SEC:
                exp = dict(t)
                exp['_iid'] = k
                # Tambah sisa off ke total_off
                if t.get('off_start'):
                    exp['total_off'] = t.get('total_off', 0.0) + (now - t['off_start'])
                expired.append(exp)

        for k in [k for k, t in self.tracks.items()
                  if t['missing'] and (now - t['last_seen']) > GRACE_SEC]:
            del self.tracks[k]

        self._rank()
        return self.tracks, expired


# ─────────────────────────────────────────────
#  EVENT LOGGER  (terminal + Excel)
# ─────────────────────────────────────────────
class EventLogger:
    def __init__(self, recorder: ExcelRecorder):
        self.rec         = recorder
        self._state      = {}
        self._did_map    = {}
        self._miss_start = {}
        self._last_tick  = {}

    def _id1_now(self, tracks):
        return next((t['display_id'] for t in tracks.values()
                     if t['display_id'] == 1), None)

    def update(self, tracks, expired_log):
        now = time.time()

        # 1. Objek baru masuk
        for iid, t in tracks.items():
            if iid not in self._did_map:
                sep()
                cp(f"  [{ts()}]  + MASUK   ID {t['display_id']}  "
                   f"conf {t['conf']:.2f}", 'gr', 'b')
                self._did_map[iid] = t['display_id']
                self._state[iid]   = 'active'
                self.rec.record_event(
                    'MASUK', t['display_id'], iid, t['conf'],
                    t['enter_time'], 0, None, False,
                    self._id1_now(tracks),
                    "Wajah pertama terdeteksi")
                self.rec.update_summary_buffer(
                    iid, t['display_id'], t['conf'],
                    t['enter_time'], 0, 0, 0, 0)

        # 2. Promosi ID
        for iid, t in tracks.items():
            prev = self._did_map.get(iid)
            if prev is not None and t['display_id'] != prev:
                sep()
                cp(f"  [{ts()}]  ^ PROMOSI ID {prev} → "
                   f"ID {t['display_id']}", 'cy', 'b')
                on_dur = (now - t['enter_time']) - t.get('total_off', 0)
                self.rec.record_event(
                    'PROMOSI', t['display_id'], iid, t['conf'],
                    t['enter_time'], on_dur, t.get('total_off', 0),
                    t.get('n_kembali', 0) > 0,
                    self._id1_now(tracks),
                    f"Naik posisi dari ID {prev} ke ID {t['display_id']}")
                self._did_map[iid] = t['display_id']

        # 3. Mulai hilang
        for iid, t in tracks.items():
            if t['missing'] and self._state.get(iid) == 'active':
                self._state[iid]      = 'missing'
                self._miss_start[iid] = now
                self._last_tick[iid]  = now
                sep()
                cp(f"  [{ts()}]  o HILANG  ID {t['display_id']}  "
                   f"grace {GRACE_SEC:.1f}s", 'yw', 'b')
                on_dur = (now - t['enter_time']) - t.get('total_off', 0)
                self.rec.record_event(
                    'HILANG', t['display_id'], iid, t['conf'],
                    t['enter_time'], on_dur, None,
                    t.get('n_kembali', 0) > 0,
                    self._id1_now(tracks),
                    f"Keluar frame — grace {GRACE_SEC:.1f}s dimulai")

        # 4. Countdown tiap 1 detik
        for iid, t in tracks.items():
            if self._state.get(iid) == 'missing':
                ms = now - self._miss_start.get(iid, now)
                if now - self._last_tick.get(iid, now) >= 1.0:
                    cp(f"         | ID {t['display_id']}  "
                       f"off {ms:.1f}s  sisa {max(0,GRACE_SEC-ms):.1f}s", 'yw')
                    self._last_tick[iid] = now

        # 5. Kembali terdeteksi
        for iid, t in tracks.items():
            if not t['missing'] and self._state.get(iid) == 'missing':
                ms = now - self._miss_start.get(iid, now)
                self._state[iid] = 'active'
                sep()
                cp(f"  [{ts()}]  < KEMBALI ID {t['display_id']}  "
                   f"off {format_dur(ms)}", 'cy', 'b')
                on_dur = (now - t['enter_time']) - t.get('total_off', 0)
                self.rec.record_event(
                    'KEMBALI', t['display_id'], iid, t['conf'],
                    t['enter_time'], on_dur, ms,
                    True, self._id1_now(tracks),
                    f"Kembali setelah {format_dur(ms)} off "
                    f"(total {t.get('n_kembali',0)}×)")

        # 6. Expire → siapa ID 1 baru
        for exp in expired_log:
            iid      = exp.get('_iid', -1)
            miss_dur = now - exp.get('last_seen', now)
            on_dur   = max(0, (now - exp['enter_time'])
                          - exp.get('total_off', 0) - miss_dur)
            sep()
            cp(f"  [{ts()}]  x EXPIRE  ID {exp['display_id']}  "
               f"on {format_dur(on_dur)}  "
               f"off total {format_dur(exp.get('total_off',0))}", 'rd', 'b')
            new_id1 = self._id1_now(tracks)
            if new_id1:
                cp(f"         L ID 1 → ID {new_id1}", 'cy', 'b')
            else:
                cp("         L Tidak ada objek aktif.", 'gy')
            self.rec.record_event(
                'EXPIRE', exp['display_id'], iid,
                exp.get('conf', 0),
                exp['enter_time'], on_dur,
                exp.get('total_off', 0),
                exp.get('n_kembali', 0) > 0,
                new_id1,
                f"Grace habis. Kembali {exp.get('n_kembali',0)}×. "
                f"Anomali {exp.get('n_anomali',0)}×")
            self.rec.mark_expired(iid)

        # 7. Update ringkasan buffer tiap frame
        for iid, t in tracks.items():
            on_dur  = (now - t['enter_time']) - t.get('total_off', 0)
            off_cur = t.get('total_off', 0)
            if t['missing'] and t.get('off_start'):
                off_cur += now - t['off_start']
            self.rec.update_summary_buffer(
                iid, t['display_id'], t['conf'],
                t['enter_time'], on_dur, off_cur,
                t.get('n_kembali', 0), t.get('n_anomali', 0))

        # Cleanup
        alive = set(tracks.keys())
        for d in [k for k in list(self._state)      if k not in alive]:
            del self._state[d]
        for d in [k for k in list(self._did_map)    if k not in alive]:
            del self._did_map[d]
        for d in [k for k in list(self._miss_start) if k not in alive]:
            del self._miss_start[d]
        for d in [k for k in list(self._last_tick)  if k not in alive]:
            del self._last_tick[d]


# ─────────────────────────────────────────────
#  ANOTASI FRAME
# ─────────────────────────────────────────────
COLORS = [(0,220,255),(0,255,120),(255,180,0),
          (200,80,255),(255,80,80),(80,180,255)]

def draw_tracks(frame, tracks):
    now = time.time()
    for _, t in tracks.items():
        did   = t['display_id']
        color = COLORS[(did-1) % len(COLORS)]
        # ── Gunakan 'bbox' (bukan 'face_bbox') ──
        fx1,fy1,fx2,fy2 = [int(v) for v in t['bbox']]
        thick   = 3 if did==1 else 2
        miss    = t.get('missing', False)
        miss_s  = (now - t['last_seen']) if miss else 0.0
        on_dur  = (now - t['enter_time']) - t.get('total_off', 0)
        if miss and t.get('off_start'):
            cur_off = now - t['off_start']
        else:
            cur_off = 0.0
        enter = time.strftime('%H:%M:%S', time.localtime(t['enter_time']))

        # Bbox: solid atau putus-putus
        if miss:
            for p1,p2 in [((fx1,fy1),(fx2,fy1)),((fx2,fy1),(fx2,fy2)),
                           ((fx2,fy2),(fx1,fy2)),((fx1,fy2),(fx1,fy1))]:
                tot=int(np.hypot(p2[0]-p1[0],p2[1]-p1[1]))
                for s in range(0,tot,16):
                    ex=int(p1[0]+(p2[0]-p1[0])*min(s+10,tot)/tot)
                    ey=int(p1[1]+(p2[1]-p1[1])*min(s+10,tot)/tot)
                    sx=int(p1[0]+(p2[0]-p1[0])*s/tot)
                    sy=int(p1[1]+(p2[1]-p1[1])*s/tot)
                    cv2.line(frame,(sx,sy),(ex,ey),color,thick)
        else:
            cv2.rectangle(frame,(fx1,fy1),(fx2,fy2),color,thick)

        # Badge ID pojok kiri atas
        id_txt = f"ID {did}"
        sc     = 0.85 if did==1 else 0.70
        (iw,ih),_ = cv2.getTextSize(id_txt, cv2.FONT_HERSHEY_DUPLEX, sc, 2)
        pad=7; bx2,by2=fx1+iw+pad*2, fy1+ih+pad*2
        cv2.rectangle(frame,(fx1,fy1),(bx2,by2),color,-1)
        if did==1:
            cv2.rectangle(frame,(fx1,fy1),(bx2,by2),(255,255,255),1)
        cv2.putText(frame,id_txt,(fx1+pad,fy1+ih+pad-1),
                    cv2.FONT_HERSHEY_DUPLEX,sc,(255,255,255),2,cv2.LINE_AA)

        # Outline tubuh semi-transparan
        px1,py1,px2,py2=[int(v) for v in t['person_bbox']]
        ov=frame.copy()
        cv2.rectangle(ov,(px1,py1),(px2,py2),color,1)
        cv2.addWeighted(ov,0.3,frame,0.7,0,frame)

        # Label DI ATAS bbox
        rows = [
            (f"Masuk : {enter}",              color,       (0,0,0)),
            (f"ON    : {format_dur(on_dur)}", (0,160,60),  (255,255,255)),
        ]
        if miss and cur_off < GRACE_SEC:
            rows.append((
                f"OFF   : {cur_off:.2f}s  sisa {max(0,GRACE_SEC-cur_off):.1f}s",
                (0,100,210),(255,255,255)))
        if t.get('n_anomali',0) > 0:
            rows.append((
                f"ANOMALI x{t['n_anomali']}",
                (0,0,180),(255,220,0)))

        for i,(lbl,bg,tc) in enumerate(reversed(rows)):
            (tw,th),_=cv2.getTextSize(lbl,cv2.FONT_HERSHEY_SIMPLEX,0.42,1)
            by=fy1-6-i*(th+5)
            if by<th+6: by=fy2+36+i*(th+6)
            cv2.rectangle(frame,(fx1,by-th-2),(fx1+tw+4,by+2),bg,-1)
            cv2.putText(frame,lbl,(fx1+2,by),
                        cv2.FONT_HERSHEY_SIMPLEX,0.42,tc,1,cv2.LINE_AA)

        cx,cy=centroid(t['bbox'])
        cv2.circle(frame,(cx,cy),4,color,-1)


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    sep()
    cp("  RT-DETR Face Tracker + Excel Recorder", 'b', 'wh')
    cp(f"  Grace={GRACE_SEC}s  "
       f"Anomali: speed>{ANOMALY_SPEED_PX_PER_SEC}px/s  "
       f"conf_drop>{ANOMALY_CONF_DROP}  "
       f"size>{ANOMALY_SIZE_RATIO}x", 'gy')
    cp("  'Q' keluar  |  'R' reset track", 'gy')
    sep()

    recorder      = ExcelRecorder()
    model         = RTDETR('rtdetr-l.pt')
    validator     = FaceValidator()
    manager       = IDManager()
    ev_logger     = EventLogger(recorder)
    anomaly_det   = AnomalyDetector()
    session_start = time.time()

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Kamera tidak ditemukan."); return

    cv2.namedWindow("Face Tracker + Excel")

    while cap.isOpened():
        ok, frame = cap.read()
        if not ok: break

        # ── Deteksi RT-DETR → filter person
        raw = list(model(frame, stream=True))
        person_dets = []
        for r in raw:
            for box in r.boxes:
                if int(box.cls[0]) != PERSON_CLASS_ID: continue
                if float(box.conf[0]) < CONF_THRESHOLD: continue
                x1,y1,x2,y2=[int(v) for v in box.xyxy[0].tolist()]
                person_dets.append({
                    'person_bbox': [x1,y1,x2,y2],
                    'conf':        float(box.conf[0]),
                })

        # ── Validasi wajah
        face_dets = []
        for pd in person_dets:
            ok_f, fb = validator.has_face(
                frame, pd['person_bbox'], min_confidence=0.5)
            if ok_f and fb:
                face_dets.append({
                    'person_bbox': pd['person_bbox'],
                    'face_bbox':   fb,
                    'conf':        pd['conf'],
                })

        # ── Update tracker
        tracks, expired = manager.update(frame, face_dets)

        # ── Deteksi anomali — gunakan t['bbox'] yang sudah konsisten
        for iid, t in tracks.items():
            if not t['missing']:
                bbox_for_anomaly = t.get('bbox')   # ← PERBAIKAN KEY
                if bbox_for_anomaly:
                    anoms = anomaly_det.check(
                        iid, t['display_id'],
                        bbox_for_anomaly,           # ← tidak ada KeyError
                        t['conf'], recorder)
                    if anoms:
                        t['n_anomali'] = t.get('n_anomali', 0) + len(anoms)

        # Cleanup anomaly detector untuk track yang expire
        for exp in expired:
            anomaly_det.remove(exp.get('_iid', -1))

        # ── Log terminal + Excel
        ev_logger.update(tracks, expired)

        # ── Gambar frame
        draw_tracks(frame, tracks)

        h, w = frame.shape[:2]
        cv2.rectangle(frame,(0,0),(w,32),(15,15,15),-1)
        cv2.putText(frame,
                    f"Wajah aktif: {len(tracks)}  |  "
                    f"Excel: {recorder.filename}",
                    (10,21), cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,(200,200,200),1,cv2.LINE_AA)

        cv2.imshow("Face Tracker + Excel", frame)
        key = cv2.waitKey(1) & 0xFF
        if key in (ord('q'), ord('Q')): break
        if key in (ord('r'), ord('R')):
            manager   = IDManager()
            ev_logger = EventLogger(recorder)
            sep(); cp("  Track di-reset.", 'yw'); sep()

    # ── Terminate
    dur = time.time() - session_start
    recorder.record_terminate(manager._id_counter, dur)
    recorder.write_summary(manager.tracks)
    cap.release()
    cv2.destroyAllWindows()
    sep()
    cp(f"  Sesi selesai. Durasi: {fmt(dur)}. "
       f"Total ID: {manager._id_counter}", 'gr', 'b')
    cp(f"  File: {recorder.filename}", 'cy', 'b')
    sep()


if __name__ == "__main__":
    main()